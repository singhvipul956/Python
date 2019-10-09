import openpyxl
from openpyxl import Workbook
import os

def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

def readData(file, sheetName, row, column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.cell(row = row, column = column).value)

def writeData(file, sheetName, row, column, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row = row, column = column).value = data
    workbook.save(file)


def createWorkBook():
    return Workbook()

def createWorkSheet(workBook, sheetName):
    newsheet = workBook.active
    newsheet.title = sheetName
    return newsheet

def saveWorkbook(fileName, filePath, workBook):
    os.chdir(filePath)
    workBook.save(filename = fileName)
