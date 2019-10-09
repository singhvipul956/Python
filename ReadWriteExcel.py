import openpyxl
import os
from openpyxl import Workbook

path = "C:\\Users\\vsing035\\Desktop\\DF\\New Hire Masterlist.xlsx"
workbook = openpyxl.load_workbook(path)
for sheetName in workbook.sheetnames:
    sheet = workbook[sheetName]
    newworkbook = Workbook()
    newsheet = newworkbook.active
    newsheet.title = sheetName
    
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            newsheet.cell(row = row, column = column).value = sheet.cell(row = row, column = column).value
    os.chdir("C:\\Users\\vsing035\\Desktop\\pytest\\Excel")
    newworkbook.save(filename = sheetName + ".xlsx")
